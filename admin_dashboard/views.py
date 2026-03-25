from django.shortcuts import render, get_object_or_404, redirect
from .models import *
from django.contrib.auth.decorators import user_passes_test
from .forms import *
from django.contrib import messages
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def update_project_order(request):
    if request.method == "POST":
        data = json.loads(request.body)
        for item in data:
            # Updating 'order' field based on drag position
            Project.objects.filter(id=item['id']).update(order=item['position'])
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'invalid request'})


def admin_login(request):
    if request.method == "POST":
        email_input = request.POST.get('email')
        password_input = request.POST.get('password')
        
        try:
            user_obj = User.objects.get(email=email_input)
            user = authenticate(request, username=user_obj.username, password=password_input)
            
            if user is not None and user.is_superuser:
                login(request, user)
                messages.success(request, f"Welcome Admin: {user.username}")
                return redirect('video')
            else:
                messages.error(request, "Invalid Password or Not a Superuser.")
                return redirect('admin_login')
                
        except User.DoesNotExist:
            messages.error(request, "This email is not registered as an Admin.")
            return redirect('admin_login')
            
    return render(request, "admin_dashboard/admin_login.html")
@user_passes_test(lambda u: u.is_superuser)
def dashboard(request):
    return render(request,"admin_dashboard/dashboard.html")


@user_passes_test(lambda u: u.is_superuser)
def video(request):
    video = Banner_video.objects.last()
    form = BannerVideoForm()
    print("=============",video)
    return render(request,'admin_dashboard/video.html',{'video':video,'form':form})

@user_passes_test(lambda u: u.is_superuser)
def change_video(request, id):
    video = get_object_or_404(Banner_video, id=id)
    if request.method == 'POST':
        form = BannerVideoForm(request.POST, request.FILES, instance=video)
        if form.is_valid():
            form.save()
            messages.success(request, "Updated successfully!")
            return redirect('video')
        else:
            messages.error(request, "Please try again.")
            return redirect('video')
        


@user_passes_test(lambda u: u.is_superuser)
def create_project(request):
    amenities = Amenity.objects.all()
    categories = LocationCategory.objects.all()

    if request.method == 'POST':
        project = Project.objects.create(
            name=request.POST.get('name'),
            project_config = request.POST.get('project_config'),
            project_type=request.POST.get('project_type'),
            project_status = request.POST.get('project_status'),
            price=request.POST.get('price'),
            location=request.POST.get('location'),
            full_location=request.POST.get('full_location'),
            description = request.POST.get('description'),
            total_units=request.POST.get('total_units'),
            sold_units=request.POST.get('sold_units'),
            construction_status=request.POST.get('construction_status'),
            project_video=request.FILES.get('project_video'),
            brochure=request.FILES.get('brochure'),
            image1=request.FILES.get('image1'),
            image2=request.FILES.get('image2'),
            image3=request.FILES.get('image3'),
            google_map_location=request.POST.get('google_map_location'),
        )

        # Amenities (M2M)
        amenity_ids = request.POST.getlist('amenities')
        project.amenities.set(amenity_ids)

        # Floor Plans (multiple)
        for img in request.FILES.getlist('floor_plans'):
            FloorPlan.objects.create(project=project, image=img)

        # Gallery (multiple)
        for img in request.FILES.getlist('gallery'):
            Gallery.objects.create(project=project, image=img)

        # Location Advantages
        categories = request.POST.getlist('adv_category')
        titles = request.POST.getlist('adv_title')
        distances = request.POST.getlist('adv_distance')

        for i in range(len(titles)):
            if titles[i]:
                LocationAdvantage.objects.create(
                    project=project,
                    category=categories[i],
                    title=titles[i],
                    distance=distances[i]
                )
        messages.success(request, "Project Saved...")
        return redirect('projects_list')

    return render(request, 'admin_dashboard/create_project.html', {'amenities': amenities,'categories':categories})


@user_passes_test(lambda u: u.is_superuser)
def edit_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    amenities = Amenity.objects.all()
    categories = LocationCategory.objects.all()  # NEW

    floor_plans = FloorPlan.objects.filter(project=project)
    gallery = Gallery.objects.filter(project=project)
    advantages = LocationAdvantage.objects.filter(project=project)

    if request.method == 'POST':

        # BASIC
        project.name = request.POST.get('name')
        project.project_config = request.POST.get('project_config')
        project.project_type = request.POST.get('project_type')
        project.project_status = request.POST.get('project_status')
        project.price = request.POST.get('price')
        project.location = request.POST.get('location')
        project.full_location = request.POST.get('full_location')
        project.description = request.POST.get('description')
        project.total_units = request.POST.get('total_units')
        project.sold_units = request.POST.get('sold_units')
        project.construction_status = request.POST.get('construction_status')
        project.google_map_location = request.POST.get('google_map_location')

        # IMAGE DELETE
        for field in ['image1', 'image2', 'image3']:
            if request.POST.get(f'delete_{field}'):
                img = getattr(project, field)
                if img:
                    img.delete(save=False)
                setattr(project, field, None)

        # FILE UPLOAD
        if request.FILES.get('project_video'):
            project.project_video = request.FILES.get('project_video')

        if request.FILES.get('brochure'):
            project.brochure = request.FILES.get('brochure')

        for field in ['image1', 'image2', 'image3']:
            if request.FILES.get(field):
                setattr(project, field, request.FILES.get(field))

        project.save()

        # AMENITIES
        project.amenities.set(request.POST.getlist('amenities'))

        # FLOOR DELETE
        delete_floor_ids = request.POST.getlist('delete_floor')
        if delete_floor_ids:
            FloorPlan.objects.filter(id__in=delete_floor_ids, project=project).delete()

        # GALLERY DELETE
        delete_gallery_ids = request.POST.getlist('delete_gallery')
        if delete_gallery_ids:
            Gallery.objects.filter(id__in=delete_gallery_ids, project=project).delete()

        # ADD NEW
        for img in request.FILES.getlist('floor_plans'):
            FloorPlan.objects.create(project=project, image=img)

        for img in request.FILES.getlist('gallery'):
            Gallery.objects.create(project=project, image=img)

        # 🔥 LOCATION ADVANTAGES (UPDATED)
        advantages.delete()

        categories_ids = request.POST.getlist('adv_category')
        titles = request.POST.getlist('adv_title')
        distances = request.POST.getlist('adv_distance')

        for i in range(len(titles)):
            if titles[i] and categories_ids[i]:
                LocationAdvantage.objects.create(
                    project=project,
                    category_id=categories_ids[i],  # ✅ FK FIX
                    title=titles[i],
                    distance=distances[i]
                )
        messages.success(request, "Project Saved...")
        return redirect('projects_list')

    return render(request, 'admin_dashboard/create_project.html', {
        'project': project,
        'amenities': amenities,
        'categories': categories,  # NEW
        'floor_plans': floor_plans,
        'gallery': gallery,
        'advantages': advantages,
    })




@user_passes_test(lambda u: u.is_superuser)
def projects_list(request):
    datas = Project.objects.all().order_by('order')
    return render(request, 'admin_dashboard/projects_list.html', {'datas': datas})


@user_passes_test(lambda u: u.is_superuser)
def delete_project(request,id):
    res = Project.objects.get(id=id)
    res.delete()
    return redirect('projects_list')



# Blogs
@user_passes_test(lambda u: u.is_superuser)
def blog_list(request):
    blog_data = BlogPost.objects.all()
    return render(request,"admin_dashboard/blog_list.html",{'blog_data':blog_data})


@user_passes_test(lambda u: u.is_superuser)
def blog_create(request):
    if request.method == "POST":
        form = BlogPostForm(request.POST, request.FILES)
        if form.is_valid():
            blog = form.save(commit=False)
            blog.title = blog.title.replace("/", ",")
            blog.save()
            messages.success(request, "✅ Blog created successfully.")
            return redirect("blog_list")
        else:
            messages.error(request, form.errors)
    else:
        form = BlogPostForm()

    return render(request, "admin_dashboard/blog_create.html", {
        "form": form,
    })


@user_passes_test(lambda u: u.is_superuser)
def blog_edit(request,id):
    blog = get_object_or_404(BlogPost,id=id)
    if request.method == "POST":
        form = BlogPostForm(request.POST,request.FILES,instance=blog)
        if form.is_valid():
            blog = form.save(commit=False)
            blog.title = blog.title.replace("/", ",")
            blog.save()
            messages.success(request,"Blog Edited...")
            return redirect("blog_list")
        else:
            messages.error(request,form.errors)
    else:
        form=BlogPostForm(instance=blog)

    return render(request,"admin_dashboard/blog_edit.html",{'form':form})


@user_passes_test(lambda u: u.is_superuser)
def blog_delete(request,id):
    res = BlogPost.objects.get(id=id)
    res.delete()
    messages.success(request, 'Blog has been deleted')
    return redirect('blog_list')





# Carrere


@user_passes_test(lambda u: u.is_superuser)
def job_list(request):
    job_data = Career.objects.all().order_by("-posted_date")
    return render(request,"admin_dashboard/job_list.html",{'job_data':job_data})


@user_passes_test(lambda u: u.is_superuser)
def create_job(request):
    if request.method == "POST":
        form = CareerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "New job added successfully.")
            return redirect("job_list")
    else:
        form = CareerForm()
    return render(request,"admin_dashboard/create_job.html", {"form": form})

@user_passes_test(lambda u: u.is_superuser)
def edit_job(request,id):
    job = get_object_or_404(Career, id=id)

    if request.method == "POST":
        form = CareerForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            messages.success(request, "Job updated successfully.")
            return redirect("job_list")
    else:
        form = CareerForm(instance=job)
    return render(request,"admin_dashboard/create_job.html",{"form":form})


@user_passes_test(lambda u: u.is_superuser)
def delete_job(request,id):
    job = get_object_or_404(Career, id=id)
    job.delete()
    messages.success(request, "Job deleted successfully.")
    return redirect("job_list")

@user_passes_test(lambda u: u.is_superuser)
def job_applications(request):
    applications = JobApplication.objects.all()
    JobApplication.objects.filter(is_seen=False).update(is_seen=True)
    return render(request,"admin_dashboard/job_applications.html",{'applications':applications})

@user_passes_test(lambda u: u.is_superuser)
def download_job_applications_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # Header Row
    headers = [
        "Job Title",
        "Full Name",
        "Email",
        "Phone",
        "Current Location",
        "Total Experience",
        "Highest Qualification",
        "Notice Period",
        "Professional Details",
        "Applied On"
    ]
    ws.append(headers)

    # Data Rows
    applications = JobApplication.objects.select_related("job").all()

    for app in applications:
        ws.append([
            app.job.title,
            app.full_name,
            app.email,
            app.phone,
            app.current_location,
            app.total_experience,
            app.highest_qualification,
            app.notice_period or "N/A",
            app.professional_details,
            app.applied_on.strftime("%d-%m-%Y %H:%M")
        ])

    # HTTP Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="job_applications.xlsx"'

    wb.save(response)
    return response







# Contact details


@user_passes_test(lambda u: u.is_superuser)
def contact_details(request):
    datas = Project_Enquiry.objects.all().order_by('-created_at')
    Project_Enquiry.objects.filter(is_seen=False).update(is_seen=True)
    return render(request, "admin_dashboard/contact_details.html", {'datas': datas})

@user_passes_test(lambda u: u.is_superuser)
def delete_contact(request, id):
    contact = get_object_or_404(Project_Enquiry, id=id)
    contact.delete()
    return redirect('contact_details')

@user_passes_test(lambda u: u.is_superuser)
def download_contacts_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    headers = ["Name", "Email", "Phone", "Project", "Message", "Created At"]
    ws.append(headers)

    contacts = Project_Enquiry.objects.all().order_by("-created_at")

    for contact in contacts:
        ws.append([
            contact.full_name,
            contact.email,
            contact.mobile_number,
            str(contact.project.name) if contact.project else "N/A",
            contact.message,
            contact.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="contact_details.xlsx"'
    wb.save(response)
    return response
@user_passes_test(lambda u: u.is_superuser)
def update_receiver_email(request):
    if request.method == "POST":
        selected_id = request.POST.get('selected_id')
        new_email = request.POST.get('new_email')

        EmailSetting.objects.all().update(is_selected=False)

        if new_email:
            obj, created = EmailSetting.objects.get_or_create(email=new_email)
            obj.is_selected = True
            obj.save()
            messages.success(request, f"Receiver email set to {new_email}")
        elif selected_id:
            obj = EmailSetting.objects.get(id=selected_id)
            obj.is_selected = True
            obj.save()
            messages.success(request, "Receiver email updated from list")

    return redirect(request.META.get('HTTP_REFERER', '/'))

#  Project Enquiry

@user_passes_test(lambda u: u.is_superuser)
def project_enquiry(request):
    enquiries = Project_Enquiry.objects.all().order_by('-created_at')
    return render(request, "admin_dashboard/project_enquiry.html", {'enquiries': enquiries})
def download_enquiries_excel(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    enquiries = Project_Enquiry.objects.all().order_by("-created_at")

    if from_date:
        enquiries = enquiries.filter(created_at__date__gte=from_date)
    if to_date:
        enquiries = enquiries.filter(created_at__date__lte=to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Project Enquiries"
    
    headers = ["Name", "Email", "Mobile", "Project", "Budget", "Message", "Date"]
    ws.append(headers)

    for e in enquiries:
        ws.append([
            e.full_name,
            e.email,
            e.mobile_number,
            str(e.project.name) if e.project else "General",
            e.budget if e.budget else "N/A",
            e.message if e.message else "N/A",
            e.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="project_enquiries.xlsx"'
    wb.save(response)
    return response
@user_passes_test(lambda u: u.is_superuser)
def edit_project_enquiry(request, pk):
    enquiry = get_object_or_404(Project_Enquiry, pk=pk)
    projects = Project.objects.all()

    if request.method == "POST":
        enquiry.full_name = request.POST.get('full_name')
        enquiry.email = request.POST.get('email')
        enquiry.mobile_number = request.POST.get('mobile_number')
        enquiry.budget = request.POST.get('budget')
        enquiry.message = request.POST.get('message')

        project_id = request.POST.get('project')
        if project_id:
            enquiry.project_id = project_id
        else:
            enquiry.project = None

        enquiry.save()

        messages.success(request, "Enquiry Saved...")
        return redirect('project_enquiry')

    return render(request, "admin_dashboard/project_enquiry.html", {
        'enquiry': enquiry,
        'projects': projects,
    })

@user_passes_test(lambda u: u.is_superuser)
def delete_project_enquiry(request, id):
    enquiry = get_object_or_404(Project_Enquiry, id=id)
    enquiry.delete()
    messages.success(request, "Enquiry deleted successfully!")
    return redirect('project_enquiry')

@user_passes_test(lambda u: u.is_superuser)
def admin_logout(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect("admin_login")